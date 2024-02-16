import os
import sqlite3
from openpyxl import load_workbook


class recup_val:
    def __init__(self, dossier):
        self.dossier = dossier

    def recuperer_valeurs(self):
        valeurs_globales = []

        # Parcourir tous les fichiers dans le dossier spécifié
        for fichier in os.listdir(self.dossier):
            if fichier.startswith("S") and fichier.endswith(".xlsx"):
                chemin_fichier = os.path.join(self.dossier, fichier)
                valeurs_fichier = []

                # Charger le fichier Excel
                wb = load_workbook(chemin_fichier, read_only=True)

                # Parcourir tous les onglets sauf le premier
                for onglet in wb.sheetnames[1:]:
                    recup = wb[onglet]

                    # Récupérer la valeur de la ligne 2, colonne 2
                    valeur_case = recup.cell(row=2, column=2).value

                    # Récupérer les valeurs spécifiques pour chaque onglet
                    valeurs_onglet = []
                    i = 56
                    while True:
                        valeur_case_1 = [recup.cell(row=i, column=1).value,
                                         recup.cell(row=i, column=12).value,
                                         recup.cell(row=i, column=13).value,
                                         recup.cell(row=i, column=14).value,
                                         recup.cell(row=i, column=15).value]
                        if any(valeur_case_1):
                            valeurs_onglet.append(valeur_case_1)
                            i += 1
                        else:
                            break

                    # Récupérer la ligne 65
                    valeur_case_2 = [recup.cell(row=65, column=1).value,
                                     recup.cell(row=65, column=12).value,
                                     recup.cell(row=65, column=13).value,
                                     recup.cell(row=65, column=14).value,
                                     recup.cell(row=65, column=15).value]
                    valeurs_onglet.append(valeur_case_2)

                    valeurs_fichier.append((fichier, onglet, valeur_case, valeurs_onglet))

                valeurs_globales.append(valeurs_fichier)

        return valeurs_globales

    def inserer_dans_bd(self, valeurs_recuperes):
        conn = sqlite3.connect('database/database.db')
        cursor = conn.cursor()

        # Réinitialiser la table en supprimant toutes les lignes
        cursor.execute("DELETE FROM FichiersGenere")

        for valeurs_fichier in valeurs_recuperes:
            for fichier, onglet, _, valeurs_onglet in valeurs_fichier:
                nom_fichier_sans_extension = os.path.splitext(fichier)[0]  # Récupérer le nom de fichier sans extension
                for ligne in valeurs_onglet:
                    # Insérer les données dans la base de données
                    cursor.execute(
                        "INSERT INTO FichiersGenere (Semestre, Ressources, Prof, H_CM, H_TD, H_TP) VALUES (?, ?, ?, ?, ?, ?)",
                        (nom_fichier_sans_extension, onglet, ligne[0], ligne[1], ligne[2], ligne[3]))

        conn.commit()
        conn.close()


# Dossier contenant les fichiers Excel
dossier_excel = 'fichiers genere'

# Créer une instance de la classe recup_val et récupérer les valeurs pour tous les fichiers
recup_val_instance = recup_val(dossier_excel)
valeurs_recuperes = recup_val_instance.recuperer_valeurs()

# Insérer les valeurs récupérées dans la base de données
recup_val_instance.inserer_dans_bd(valeurs_recuperes)


# Afficher les valeurs récupérées pour chaque fichier
for valeurs_fichier in valeurs_recuperes:
    for fichier, onglet, valeur_case, valeurs_onglet in valeurs_fichier:
        print(f"Nom du fichier : {os.path.splitext(fichier)[0]}")
        print(f"Onglet : {onglet}")
        print("Valeur de la ligne 2, colonne 2 :", valeur_case)
        for ligne in valeurs_onglet:
            print("Valeurs de la ligne:", ligne)
        print()
