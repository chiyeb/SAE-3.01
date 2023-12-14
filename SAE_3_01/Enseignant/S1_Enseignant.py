import sqlite3
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.copier import WorksheetCopy

# Connexion à la base de données SQLite
try:
    with sqlite3.connect("../database/database.db") as conn:
        cursor = conn.cursor()

        # Récupérer les noms uniques des semestres
        cursor.execute("SELECT DISTINCT Semestre FROM Planning WHERE Semestre = 'S1'")
        semesters = [row[0] for row in cursor.fetchall()]

        # Ouvrir le fichier "fichier_vierge.xlsx" pour obtenir les paramètres
        try:
            fichier_vierge = openpyxl.load_workbook("fichier_vierge.xlsx", read_only=False)
        except FileNotFoundError:
            print("Le fichier fichier_vierge.xlsx n'a pas été trouvé.")
            exit()

        for semester in semesters:
            # Exécutez une requête SQL pour récupérer toutes les ressources pour le semestre actuel
            cursor.execute("SELECT DISTINCT Ressource FROM Planning WHERE Semestre = ?", (semester,))

            # Récupérez toutes les ressources dans une liste
            resources = [row[0] for row in cursor.fetchall()]

            for resource in resources:
                # Ouvrir le fichier vierge pour obtenir les paramètres
                base_sheet = fichier_vierge["Feuil1"]

                # Créer une nouvelle feuille avec le nom de la ressource
                worksheet = fichier_vierge.copy_worksheet(base_sheet)
                worksheet.title = resource

                # Exécutez une requête SQL pour récupérer les données de la ressource actuelle
                cursor.execute("SELECT Ressource, H_CM, H_TD, H_TP, Resp FROM Planning WHERE Ressource = ? AND Semestre = ?", (resource, semester))

                # Récupérez les données dans une liste de tuples
                data_from_database = cursor.fetchall()

                # Écrire les données spécifiques dans la nouvelle feuille
                if data_from_database:
                    worksheet.cell(row=5, column=2, value=data_from_database[0][1])
                    worksheet.cell(row=5, column=3, value=data_from_database[0][2])
                    worksheet.cell(row=5, column=4, value=data_from_database[0][3])
                    worksheet.cell(row=2, column=2, value=resource)
                    worksheet.cell(row=2, column=7, value=data_from_database[0][4])

        # Sauvegarder le classeur Excel
        fichier_vierge.save("S1.xlsx")

        # Message pour confirmer les ajouts dans le fichier Excel
        print("Les données ont été ajoutées à S1.xlsx")

# Si une erreur survient lors de la connexion à la base de données, affichez un message d'erreur
except sqlite3.Error as e:
    print(f"Erreur lors de la connexion à la base de données : {e}")
