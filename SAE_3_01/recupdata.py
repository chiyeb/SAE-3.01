import sqlite3
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.reader.excel import load_workbook

from insertdata import InsertData
from selectfile import *


class RecupData:
    """
    Classe permettant de récupérer certaines données dans les fichiers
    """
    instance = None
    files = None
    dossier_fichier_genere = "fichiers genere"

    def __new__(cls):
        if cls.instance is None:
            cls.instance = super(RecupData, cls).__new__(cls)
            cls.instance._setup()
        return cls.instance

    def _setup(self):
        """
        "Setup" l'objet : initialise la connexion à la BD
        :return:
        """
        self.files = SelectFile()
        # Initialise la connexion à la base de données
        self.conn = sqlite3.connect('database/database.db')
        self.cursor = self.conn.cursor()

    def trouver_val(self, semestre, semestre_onglet):
        """
        Fonction qui trouve les ressources dans le fichier planning, récupère les valeurs des heures de CM,TD et TP.
        :param semestre:
        :param semestre_onglet:
        :return:
        """
        planning = pd.ExcelFile(self.files.planning_file)
        self.cursor.execute("SELECT Libelle, Num_Res FROM Maquette WHERE Semestre = ?", (semestre,))
        resultats = self.cursor.fetchall()
        for row in resultats:
            num_res = row[1]
            libelle = row[0]
            s = pd.read_excel(planning, semestre_onglet)
            for index, row in s.iterrows():
                if num_res in row.values:
                    index_num_res = row.values.tolist().index(num_res)
                    valeur_case_3 = 0 if pd.isna(row.iloc[index_num_res + 3]) else row.iloc[index_num_res + 3]
                    valeur_case_5 = 0 if pd.isna(row.iloc[index_num_res + 5]) else row.iloc[index_num_res + 5]
                    valeur_case_7 = 0 if pd.isna(row.iloc[index_num_res + 7]) else row.iloc[index_num_res + 7]
                    valeur_case_10 = row.iloc[index_num_res + 10]
                    if (isinstance(valeur_case_3, int) and isinstance(valeur_case_7, int)
                            and isinstance(valeur_case_5, int)):
                        additions = valeur_case_3 + valeur_case_5 + valeur_case_7
                        if index_num_res + 10 < len(row) and additions > 0:
                            insertdata = InsertData()
                            insertdata.insert_planning(semestre, libelle, valeur_case_3, valeur_case_5, valeur_case_7,
                                                       valeur_case_10)

    def onglet_existe(self, fichier, nom_onglet):
        """
        Vérifie si l'onglet existe
        :param fichier: Chemin du fichier
        :param nom_onglet: Nom de l'onglet
        :return:
        """
        wb = openpyxl.load_workbook(fichier, read_only=True)
        # Retourne True si l'onglet existe, False sinon
        return nom_onglet in wb.sheetnames

    def recup_h_prof(self, semestre, semestre_onglet):
        """
        Récupère le nombre de groupes de chaque professeur
        :param semestre: Semestre
        :param semestre_onglet: Semestre de l'onglet
        :return:
        """
        # Charge le fichier Excel
        fichier = self.files.QFQ_file
        if self.onglet_existe(fichier, semestre_onglet):
            df = pd.read_excel(fichier, semestre_onglet)
            donnees = {}
            ressource_actuelle = None

            for _, row in df.iterrows():
                resource = row['Ressource']
                intervenant = row['Intervenants']
                cm = row['CM']
                td = row['TD']
                tp_non_dedoubles = row['TP (non dédoublés)']
                tp_dedoubles = row['TP (dédoublés)']
                test = row['Test']
                # Vérifie si une nouvelle ressource commence
                if pd.notna(resource):
                    ressource_actuelle = resource
                    donnees[ressource_actuelle] = {}

                # Vérifie si la ligne contient des données d'intervenant
                if pd.notna(intervenant):
                    if ressource_actuelle not in donnees:
                        donnees[ressource_actuelle] = {}
                    if intervenant not in donnees[ressource_actuelle]:
                        donnees[ressource_actuelle][intervenant] = []

                    donnees[ressource_actuelle][intervenant].append({
                        'CM': cm,
                        'TD': td,
                        'TP (non dédoublés)': tp_non_dedoubles,
                        'TP (dédoublés)': tp_dedoubles,
                        'Test': test})
            # Insère ou met à jour les enregistrements dans la base de données
            for resource, intervenant_data in donnees.items():
                if resource not in donnees:
                    donnees[resource] = {None: [{}]}
                for intervenant, data_list in intervenant_data.items():
                    if data_list:
                        for d in data_list:
                            # Vérifie si l'enregistrement existe déjà dans la base de données
                            self.cursor.execute(
                                "SELECT * FROM HoraireProf WHERE Semestre = ? AND Ressource = ? AND Intervenant = ?",
                                (semestre, resource, intervenant))
                            existing_record = self.cursor.fetchone()

                            if existing_record:
                                # Met à jour l'enregistrement existant
                                self.cursor.execute(
                                    "UPDATE HoraireProf SET CM = ?, TD = ?, TP_Non_Dedoubles = ?, TP_Dedoubles = ?, "
                                    "Test = ? WHERE Semestre = ? AND Ressource = ? AND Intervenant = ?",
                                    (d['CM'], d['TD'], d['TP (non dédoublés)'], d['TP (dédoublés)'], d['Test'],
                                     semestre, resource, intervenant))
                            else:
                                # Insère un nouvel enregistrement
                                self.cursor.execute(
                                    "INSERT INTO HoraireProf (Semestre, Ressource, Intervenant, CM, TD, "
                                    "TP_non_dedoubles, TP_Dedoubles, Test) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                                    (semestre, resource, intervenant, d['CM'], d['TD'], d['TP (non dédoublés)'],
                                     d['TP (dédoublés)'],
                                     d['Test']))

                            self.conn.commit()
                    else:
                        print("Aucune données pour cette ressource.")
                        print("\n")

    def recup_nom_prof(self):
        """
        Récupère les noms des professeurs dans le fichier texte, ainsi que leur acronymes et emails, et les insère
        ou met à jour dans la base de données.
        """
        # Ouvre le fichier en mode lecture
        with open(self.files.nom_prof_file, "r") as fichier:
            # Lire chaque ligne du fichier
            for ligne in fichier:
                # Divise la ligne en utilisant le signe "=" comme séparateur (acronyme=nom_professeur=email)
                parties = ligne.strip().split("=")
                print(ligne)
                print(parties)
                # S'assurer qu'il y a au moins un acronyme et un nom avant de continuer
                if len(parties) >= 2:
                    # Prépare les variables pour acronyme, nom et initialise email à None
                    acronyme, nom = parties[0], parties[1]
                    email = parties[2] if len(parties) == 3 else None  # Email si présent

                    # Exécute la requête SQL pour vérifier si le prof existe déjà
                    resultat_requete = self.cursor.execute("SELECT Acronyme FROM Prof WHERE Acronyme = ?",
                                                           (acronyme,)).fetchone()

                    # Vérifie si la requête a renvoyé des résultats, pour savoir si le prof existe déjà
                    if resultat_requete:
                        # Mise à jour du nom du prof et potentiellement de l'email
                        self.cursor.execute("UPDATE Prof SET NomProf = ?, MailProf = ? WHERE Acronyme = ?",
                                            (nom, email, acronyme))
                    else:
                        # Insertion du nouveau prof avec potentiellement un email
                        self.cursor.execute("INSERT INTO Prof (Acronyme, NomProf, MailProf) VALUES (?, ?, ?)",
                                            (acronyme, nom, email))
                    # Sauvegarder les changements dans la base de données
                    self.conn.commit()
                else:
                    print(f"Erreur de format ou ligne incomplète : {ligne}")

    def recup_res_couleur(self, semestre, semestre_onglet):
        """
        Fonction pour récupérer les couleurs liée à chaque ressource pour un semestre précis
        :param semestre: Semestre
        :param semestre_onglet: Onglet du semestre
        :return:
        """
        print("Récupération des couleurs")
        ressource_couleur = {}
        self.cursor.execute("SELECT Num_Res FROM Maquette WHERE Semestre = ?", (semestre,))
        resultats = [item[0] for item in self.cursor.fetchall()]
        # On ouvre le fichier excel
        fichier = openpyxl.load_workbook(self.files.planning_file)
        fichier_onglet_semestre = fichier[semestre_onglet]
        for row in fichier_onglet_semestre.iter_rows():
            for cell in row:
                # Si ressource trouvé
                if cell.value in resultats:
                    couleur = cell.fill.start_color.rgb
                    # Si couleur est autre que blanche
                    if couleur != '00000000' and couleur:
                        # On récupère la couleur de la cellule
                        ressource_couleur[cell.value] = couleur
        return ressource_couleur

    def trouver_type_cours_2e_range(self, semestre_onglet):
        """
        Récupère la deuxième occurrence de chaque type de cours dans le fichier planning
        :param semestre_onglet: Onglet du semestre
        :return:
        """
        fichier = openpyxl.load_workbook(self.files.planning_file, data_only=True)
        fichierOngletSemestre = fichier[semestre_onglet]
        type_cours_dict = {}
        for row in fichierOngletSemestre.iter_rows(min_row=1, max_row=2):
            # On itère chaque cellule
            for cell in row:
                # Si la valeur de la celulle = ["Cours", "TD", "TP", "Test"]
                if cell.value in ["Cours", "TD", "TP", "Test"]:
                    # On récupère la valeur de la colone de la cellule
                    type_cours_dict[cell.value] = cell.column

        return type_cours_dict

    def trouver_type_cours_1er_range(self, semestre_onglet):
        """
        Récupère la première occurrence de chaque type de cours dans le fichier planning
        :param semestre_onglet: Onglet du semestre
        :return:
        """
        fichier = openpyxl.load_workbook(self.files.planning_file, data_only=True)
        fichier_onglet_semestre = fichier[semestre_onglet]
        type_cours_dict = {}
        # Les valeurs à trouver
        valeurs_a_trouver = {"Cours", "TD", "TP", "Test"}
        for row in fichier_onglet_semestre.iter_rows(min_row=1, max_row=2):
            # On itère chaque cellule
            for cell in row:
                # Si la valeur de la celulle = ["Cours", "TD", "TP", "Test"]
                if cell.value in ["Cours", "TD", "TP", "Test"]:
                    # On récupère la valeur de la colonne de la cellule
                    type_cours_dict[cell.value] = cell.column
                    # On enlève de "valeur_a_trouver" la valeur de la cellule
                    valeurs_a_trouver.remove(cell.value)
                # S'il y à plus de valeur dans "valeur_a_trouver"
                if not valeurs_a_trouver:
                    return type_cours_dict
        return type_cours_dict

    def recup_X_et_Y(self, semestre, semestre_onglet):
        """
        Récupère chaque cours dans le fichier planning
        :param semestre: Semestre
        :param semestre_onglet: Onglet du semestre
        :return:
        """
        # On réinitialise les valeurs de la base de donnée
        self.cursor.execute("UPDATE Horaires SET NbCours = 0 WHERE Semestre = ?",
                            (semestre,))
        # Récupération du fichier planning
        fichier = openpyxl.load_workbook(self.files.planning_file, data_only=True)
        # Appel des fonctions nécéssaire
        ressource_couleur = self.recup_res_couleur(semestre, semestre_onglet)
        fichier_onglet_semestre = fichier[semestre_onglet]
        type_cours_dict1 = self.trouver_type_cours_1er_range(semestre_onglet)
        type_cours_dict2 = self.trouver_type_cours_2e_range(semestre_onglet)
        for col in fichier_onglet_semestre.iter_cols():
            if col[0].value is not None and 'Date' in col[0].value:
                # Parcourir les lignes pour la colonne de date
                for cell in col:
                    cell_value = cell.value
                    # Si c'est un date
                    if isinstance(cell_value, datetime):
                        date = cell_value.date()
                    else:
                        date = cell_value
                    # s'il y a une date trouvée
                    if date:
                        row_index = cell.row
                        for row_cell in fichier_onglet_semestre[row_index]:
                            # Récupérer la valeur et la couleur de chaque cellule de la ligne
                            if row_cell.value == "X" or row_cell.value == "Y":
                                if row_cell.value == "X":
                                    salle = "TD/Amphi"
                                else:
                                    salle = "Machine"
                                # Appel de la fonction typeCours qui permet de vérifier quel type de cours est une case
                                tCours = self.type_cours(type_cours_dict1, type_cours_dict2, row_cell.column)
                                valeur = row_cell.value
                                # Récupération de la couleur de la cellule
                                couleur = row_cell.fill.start_color.rgb
                                # On récupère les coordonnées du cours (X,Y)
                                idCours = row_cell.coordinate
                                for cle, valeur in ressource_couleur.items():
                                    # Si la couleur de la cellule = la couleur d'une ressource
                                    if valeur == couleur:
                                        # On vérifie si le cours existe déjà dans la BD
                                        self.cursor.execute("SELECT IdCours, Semestre FROM Cours WHERE IdCours = ? "
                                                            "AND Semestre = ?", (idCours, semestre))
                                        resultCours = self.cursor.fetchone()
                                        # On vérifie si la ressource et le type de cours existe déjà dans la BD
                                        self.cursor.execute(
                                            "SELECT Ressource, Type_Cours FROM Horaires WHERE Ressource "
                                            "= ? AND Type_Cours = ?",
                                            (cle, tCours))
                                        resultHoraires = self.cursor.fetchone()
                                        # Si le cours n'existe pas
                                        if not resultCours:
                                            # Si un commentaire existe pour la cellule
                                            if row_cell.comment:
                                                # On récupère le commentaire et l'insère
                                                commentaire = row_cell.comment.text
                                                self.cursor.execute("INSERT INTO Cours (IdCours, Semestre, Ressource, "
                                                                    "Date_Cours, Commentaire, Type_Cours, Salle) VALUES (?, "
                                                                    "?, ?, ?, ?, ?, ?)",
                                                                    (idCours, semestre, cle, date, commentaire, tCours,
                                                                     salle))
                                            # Si pas de commentaires dans la cellule
                                            else:
                                                self.cursor.execute("INSERT INTO Cours (IdCours, Semestre, Ressource, "
                                                                    "Date_Cours, Type_Cours, Salle) VALUES (?, ?, "
                                                                    "?, ?, ?, ?)",
                                                                    (idCours, semestre, cle, date, tCours, salle))
                                            self.conn.commit()
                                        # Si le cours n'existe pas
                                        else:
                                            if row_cell.comment:
                                                commentaire = row_cell.comment.text
                                                self.cursor.execute("UPDATE Cours SET Ressource = ?, Date_Cours = ?, "
                                                                    "Commentaire = ?, Type_Cours = ?, Salle = ? WHERE IdCours = ? ",
                                                                    (cle, date, commentaire, tCours, salle, idCours,))
                                            else:
                                                self.cursor.execute("UPDATE Cours SET Ressource = ?, Date_Cours = ?, "
                                                                    "Type_Cours = ?, Salle = ? WHERE IdCours = ? ",
                                                                    (cle, date, tCours, salle, idCours))
                                            self.conn.commit()
                                        # Si le type de cours existe déjà pour le semestre et la ressource dans la BD
                                        if resultHoraires:
                                            self.cursor.execute("UPDATE Horaires SET NbCours = NbCours+2 WHERE "
                                                                "Ressource = ? AND Type_Cours = ?",
                                                                (cle, tCours))
                                        else:
                                            self.cursor.execute("INSERT INTO Horaires (Semestre, Ressource, "
                                                                "Type_Cours, nbCours, Salle) VALUES (?, ?, ?, 2, ?)",
                                                                (semestre, cle, tCours, salle))
                                        self.cursor.connection.commit()

    def type_cours(self, type_cours_dict1, type_cours_dict2, col):
        """
        Fonction qui vérifie pour une case précise, quel type de cours c'est
        :param type_cours_dict1: Dictionnaire des types de cours
        :param type_cours_dict2: Dictionnaire des types de cours
        :param col: Colonne
        :return: Type de cours
        """
        t_cours = None
        if type_cours_dict1["Cours"] <= col <= type_cours_dict1["TD"]:
            t_cours = "Amphi"
        if type_cours_dict1["TD"] <= col <= type_cours_dict1["TP"]:
            t_cours = "TD"
        if type_cours_dict1["TP"] <= col < type_cours_dict1["Test"]:
            t_cours = "TP"
        if type_cours_dict1["Test"] <= col <= type_cours_dict2["Cours"]:
            t_cours = "Test"
        if type_cours_dict2["Cours"] <= col <= type_cours_dict2["TD"]:
            t_cours = "Amphi"
        if type_cours_dict2["TD"] <= col <= type_cours_dict2["TP"]:
            t_cours = "TD"
        if type_cours_dict2["TP"] <= col < type_cours_dict2["Test"]:
            t_cours = "TP"
        if col >= type_cours_dict2["Test"]:
            t_cours = "Test"
        return t_cours


