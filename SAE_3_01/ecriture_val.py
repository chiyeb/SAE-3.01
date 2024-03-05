import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from selectFile import selectFile

class ecriture_val:
    instance = None
    files = None

    def __new__(cls):
        if cls.instance is None:
            cls.instance = super(ecriture_val, cls).__new__(cls)
            cls.instance._setup()
        return cls.instance

    def __init__(self):
        # Définir les styles de bordure pour le tableau
        self.thick_border = Border(left=Side(style='thick'),
                                   right=Side(style='thick'),
                                   top=Side(style='thick'),
                                   bottom=Side(style='thick'))
        self.thin_border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

    def _setup(self):
        self.files = selectFile()
        # Initialise la connexion à la base de données
        self.conn = sqlite3.connect('database/database.db')
        self.cursor = self.conn.cursor()

    def write_data_to_excel(self, excel_file_path):
        # Charger le classeur Excel existant
        wb = load_workbook(excel_file_path)
        worksheet = wb.active

        # Vérifier si le fichier Excel est vide
        is_empty = worksheet.max_row == 1 and worksheet.max_column == 1

        # Supprimer le contenu existant du fichier Excel
        if not is_empty:
            worksheet.delete_rows(2, worksheet.max_row)

        # Récupérer les données à partir de la base de données
        self.cursor.execute("SELECT Prof, Ressources, H_CM, H_TD, H_TP, Test FROM FichiersGenere")
        data_from_database = self.cursor.fetchall()

        if data_from_database:
            print(f"Traitement des données: {data_from_database}")

            # Déterminer le nombre de colonnes dans la base de données
            num_columns = len(data_from_database[0])

            last_prof = None  # Pour suivre le dernier professeur ajouté
            row_index = 2 if is_empty else worksheet.max_row + 1  # Commencer à la deuxième ligne ou à la ligne suivante

            for row_data in data_from_database:
                prof, ressource, h_cm, h_td, h_tp, test = row_data

                # Vérifier si la ressource a un professeur associé
                if prof:
                    # Insérer une ligne vide si le professeur actuel est différent du dernier professeur ajouté
                    if last_prof != prof and not is_empty:
                        worksheet.insert_rows(row_index)
                        last_prof = prof
                        row_index += 2  # Avancer à la nouvelle ligne insérée


                    # Écrire les données dans les cellules de la ligne 1 des colonnes A à G
                    data_to_write = ["NOM DU PROF", "RESSOURCES", "NOMBRE D'HEURE DE CM", "NOMBRE D'HEURE DE TD", "NOMBRE D'HEURE DE TP", "NOMBRE D'HEURE DE TEST", "NOMBRE D'HEURE TOTAL"]
                    for index, value in enumerate(data_to_write, start=1):
                        cell = worksheet.cell(row=1, column=index, value=value).border = self.thin_border

                    # Écrire les données dans les colonnes correspondantes et appliquer les styles de bordure
                    cell_data = "" if prof is None else prof
                    cell = worksheet.cell(row=row_index, column=1, value=cell_data)
                    cell.border = self.thin_border  # Appliquer une bordure fine à la cellule

                    cell_data = "" if ressource is None else ressource
                    cell = worksheet.cell(row=row_index, column=2, value=cell_data)
                    cell.border = self.thin_border  # Appliquer une bordure fine à la cellule

                    cell_data = h_cm if h_cm is not None else 0
                    cell = worksheet.cell(row=row_index, column=3, value=cell_data)
                    cell.border = self.thin_border  # Appliquer une bordure fine à la cellule

                    cell_data = h_td if h_td is not None else 0
                    cell = worksheet.cell(row=row_index, column=4, value=cell_data)
                    cell.border = self.thin_border  # Appliquer une bordure fine à la cellule

                    cell_data = h_tp if h_tp is not None else 0
                    cell = worksheet.cell(row=row_index, column=5, value=cell_data)
                    cell.border = self.thin_border  # Appliquer une bordure fine à la cellule

                    cell_data = "" if test is None else test
                    cell = worksheet.cell(row=row_index, column=6, value=cell_data)
                    cell.border = self.thin_border  # Appliquer une bordure fine à la cellule

                    total_hours = sum(filter(None, [h_cm, h_td, h_tp]))
                    worksheet.cell(row=row_index, column=num_columns + 1, value=total_hours).border = self.thin_border

                    # Appliquer une bordure épaisse à la première ligne du groupe de professeurs
                    if last_prof != prof:
                        for col in worksheet.iter_cols(min_col=1, max_col=num_columns, min_row=row_index,
                                                       max_row=row_index):
                            for cell in col:
                                cell.border = self.thick_border

                    row_index += 1  # Avancer à la ligne suivante

            # Sauvegarder le fichier Excel
            wb.save(excel_file_path)


Ecriture_val = ecriture_val()
Ecriture_val.write_data_to_excel("fichiers genere/Professeurs_Horaires.xlsx")
