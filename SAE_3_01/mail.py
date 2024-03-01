import glob
import os
from email.mime.text import MIMEText

import openpyxl
import sqlite3
from openpyxl.workbook import Workbook
from sqlalchemy import create_engine
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import smtplib
from collections import defaultdict

print("Début du programme\nVeuillez patienter...")

# Connexion à la base de données
db = 'sqlite:///database/database.db'
table_prof_bd = 'Prof'

# Paramètres de l'e-mail ( à changer par le mail que vous voulez, actuellemnt c'est le mail de Tymeo )
email_host = 'smtp.gmail.com'
email_port = 587
email_user = 'rugierotymeo@gmail.com'
email_password = 'igtd loyc dyff suxj'

chemin_bd = 'database/database.db'


def run():
    print("Recherche des fichiers Excel...")
    fichier_excel = trouver_fichier_excel('fichiers genere')
    print("Fichiers trouvés\nTraitement des fichiers Excel...")
    traitement_fichier_excel(fichier_excel)


def verif_bd_connect():
    try:
        conn = sqlite3.connect(chemin_bd)
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"Erreur de connexion à la base de données SQLite : {e}")
        return False


def recup_mail_prof(nom_prof):
    """
    Fonction qui récupère le mail du prof en paramètre
    :param nom_prof: Nom du professeur
    """
    try:
        conn = sqlite3.connect(chemin_bd)
        cur = conn.cursor()
        cur.execute(f"SELECT MailProf FROM {table_prof_bd} WHERE NomProf = ?", (nom_prof,))
        result = cur.fetchone()
        conn.close()
        if result:
            return result[0]
        else:
            print(f"Aucun e-mail trouvé pour {nom_prof}")
            return None
    except sqlite3.Error as e:
        print(f"Une erreur est survenue lors de la récupération de l'email : {e}")
        return None


def envoie_mail(server, receiver_email, file_path):
    """
    Fonction qui envoie un mail avec le fichier en paramètre au mail du professeur en paramètre ainsi que le
    "serveur" qui envoie le mail
    :param server: Serveur SMTP
    :param receiver_email: Email du professeur
    :param file_path: Chemin du fichier à envoyer
    """
    try:
        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['To'] = receiver_email
        msg['Subject'] = 'Votre fichier Excel'

        # Changer le message du mail si besoin
        body = "Voici en pièce jointe le fichier généré automatiquement par le programme."
        msg.attach(MIMEText(body, 'plain'))

        part = MIMEBase('application', "octet-stream")
        with open(file_path, 'rb') as file:
            part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(file_path)}"')
        msg.attach(part)  # Attachez le fichier

        server.send_message(msg)
        print(f"Email envoyé à {receiver_email}")
    except Exception as e:
        print(f"Erreur lors de l'envoi de l'email : {e}")


def sauvegarde_nouvelle_feuille(chemin_fichier, nom_feuille, nouvelle_destination):
    """
    Fonction qui sauvegarde une nouvelle feuille dans un nouveau fichier
    :param chemin_fichier: Chemin du fichier
    :param nom_feuille: Nom de la feuille
    :param nouvelle_destination: Nouvelle destination du fichier
    """
    wb = openpyxl.load_workbook(chemin_fichier)
    feuille_base = wb[nom_feuille]

    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.title = nom_feuille

    for row in feuille_base.iter_rows(values_only=True):
        new_sheet.append(row)

    new_wb.save(nouvelle_destination)


def traitement_fichier_excel(chemin, dossier="fichiers genere mail"):
    """
    Fonction qui traite les fichiers Excel générés par le programme
    :param chemin: Chemin des fichiers
    """
    if not verif_bd_connect():
        print("Impossible de se connecter à la base de données. Vérifiez la configuration.")
        return

    engine = create_engine(db)
    feuille_prof = defaultdict(list)
    for chemin_fichier in chemin:
        wb = openpyxl.load_workbook(chemin_fichier)
        nom_fichier_base = os.path.splitext(os.path.basename(chemin_fichier))[0]  # Extrait le nom de base sans l'extension
        print(f"Traitement du fichier {chemin_fichier}...")
        for nom_feuille in wb.sheetnames:
            feuille = wb[nom_feuille]
            nom_prof = feuille['G2'].value
            if nom_prof:
                feuille_prof[nom_prof].append((wb, nom_feuille, nom_fichier_base))

    for nom_prof, feuilles in feuille_prof.items():
        print(f"Traitement des feuilles pour {nom_prof}...")
        new_wb = Workbook()
        new_wb.remove(new_wb.active)  # Supprime la feuille par défaut créée avec le nouveau classeur
        for wb, nom_feuille, nom_fichier_base in feuilles:
            original_sheet = wb[nom_feuille]
            # Utilise le nom du fichier dans le titre de la nouvelle feuille
            titre_nouvelle_feuille = f"{nom_fichier_base}_{nom_feuille}"
            nouvelle_feuille = new_wb.create_sheet(title=titre_nouvelle_feuille)
            for row in original_sheet.iter_rows(values_only=True):
                nouvelle_feuille.append(row)

        nom_nv_fichier = f"{nom_prof.replace(' ', '_')}_sheets.xlsx"
        chemin_nv_fichier = os.path.join(dossier, nom_nv_fichier)
        try:
            new_wb.save(chemin_nv_fichier)
            print(f"Fichier {nom_nv_fichier} sauvegardé")
        except Exception as e:
            print(f"Erreur lors de la sauvegarde du fichier {nom_nv_fichier} : {e}")

        receveur_mail = recup_mail_prof(nom_prof)
        if receveur_mail:
            with smtplib.SMTP(email_host, email_port) as server:
                server.starttls()
                server.login(email_user, email_password)
                envoie_mail(server, receveur_mail, chemin_nv_fichier)
        else:
            print(f"Email non trouvé pour {nom_prof}")


def trouver_fichier_excel(dossier):
    """
    Fonction qui trouve les fichiers Excel dans un dossier en paramètre
    :param dossier: Dossier où chercher les fichiers
    """
    return glob.glob(f"{dossier}/*.xlsx")

if __name__ == "__main__":
    run()
    print("Fin du programme")