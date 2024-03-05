import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.workbook import Workbook

from selectfile import *


class ScribeFileProf:
    """
    Classe qui écrit les données de la base de données à partir des fichiers générés dans un fichier Excel
    """
    instance = None
    files = None
    dossier = "fichiers genere"

    def __new__(cls):
        if cls.instance is None:
            cls.instance = super(ScribeFileProf, cls).__new__(cls)
            cls.instance._setup()
        return cls.instance

    def __init__(self):
        # Défini les styles de bordure pour le tableau (épaisseur de la bordure) dans le fichier Excel
        self.thick_border = Border(left=Side(style='thick'),
                                   right=Side(style='thick'),
                                   top=Side(style='thick'),
                                   bottom=Side(style='thick'))
        self.thin_border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

    def _setup(self):
        self.files = SelectFile()
        self.conn = sqlite3.connect('database/database.db')
        self.cursor = self.conn.cursor()

    def run(self):
        print("Début de l'écriture des données dans le fichier Excel...")
        self.inserer_dans_bd(self.recuperer_valeurs())
        self.write_data_to_excel("fichiers genere/Professeurs_Horaires.xlsx")

    def write_data_to_excel(self, excel_file_path):
        """
        Fonction qui écrit les données de la base de données à partir des fichiers générés dans un fichier Excel
        :param excel_file_path: Chemin du fichier Excel
        """
        print(f"Écriture des données dans le fichier Excel {excel_file_path}...")
        # Vérifie si le fichier Excel existe, sinon le crée
        if not os.path.exists(excel_file_path):
            os.makedirs(os.path.dirname(excel_file_path), exist_ok=True)  # Crée le dossier s'il n'existe pas
            wb = Workbook()
            ws = wb.active
            ws.title = "Horaires"
            wb.save(excel_file_path)
            print(f"Le fichier '{excel_file_path}' a été créé car il n'existait pas.")
        else:
            wb = load_workbook(excel_file_path)
            ws = wb.active

        wb = load_workbook(excel_file_path)
        worksheet = wb.active
        # Vérifie si le fichier Excel est vide
        is_empty = worksheet.max_row == 1 and worksheet.max_column == 1
        if not is_empty:
            worksheet.delete_rows(2, worksheet.max_row)
        # Récupère les données à partir de la base de données
        self.cursor.execute("SELECT Prof, Ressources, H_CM, H_TD, H_TP, Test FROM FichiersGenere")
        data_from_database = self.cursor.fetchall()

        if data_from_database:
            print(f"Traitement des données: {data_from_database}")

            # On détermine le nombre de colonnes dans la BD
            num_columns = len(data_from_database[0])

            last_prof = None
            row_index = 2 if is_empty else worksheet.max_row + 1

            for row_data in data_from_database:
                prof, ressource, h_cm, h_td, h_tp, test = row_data

                # On vérifie si la ressource a un professeur associé
                if prof:
                    # On insère une ligne vide si le professeur actuel est différent du dernier professeur ajouté
                    if last_prof != prof and not is_empty:
                        worksheet.insert_rows(row_index)
                        last_prof = prof
                        row_index += 2

                    # Écrire les données dans les cellules de la ligne 1 des colonnes A à G
                    data_to_write = ["NOM DU PROF", "RESSOURCES", "NOMBRE D'HEURE DE CM", "NOMBRE D'HEURE DE TD",
                                     "NOMBRE D'HEURE DE TP", "", "NOMBRE D'HEURE TOTAL"]
                    for index, value in enumerate(data_to_write, start=1):
                        cell = worksheet.cell(row=1, column=index, value=value).border = self.thin_border

                    cell_data = "" if prof is None else prof
                    cell = worksheet.cell(row=row_index, column=1, value=cell_data)
                    cell.border = self.thin_border

                    cell_data = "" if ressource is None else ressource
                    cell = worksheet.cell(row=row_index, column=2, value=cell_data)
                    cell.border = self.thin_border

                    cell_data = h_cm if h_cm is not None else 0
                    cell = worksheet.cell(row=row_index, column=3, value=cell_data)
                    cell.border = self.thin_border

                    cell_data = h_td if h_td is not None else 0
                    cell = worksheet.cell(row=row_index, column=4, value=cell_data)
                    cell.border = self.thin_border

                    cell_data = h_tp if h_tp is not None else 0
                    cell = worksheet.cell(row=row_index, column=5, value=cell_data)
                    cell.border = self.thin_border

                    cell_data = "" if test is None else test
                    cell = worksheet.cell(row=row_index, column=6, value=cell_data)
                    cell.border = self.thin_border

                    total_hours = sum(filter(None, [h_cm, h_td, h_tp]))
                    worksheet.cell(row=row_index, column=num_columns + 1, value=total_hours).border = self.thin_border

                    if last_prof != prof:
                        for col in worksheet.iter_cols(min_col=1, max_col=num_columns, min_row=row_index,
                                                       max_row=row_index):
                            for cell in col:
                                cell.border = self.thick_border

                    row_index += 1

            wb.save(excel_file_path)
            print(f"Données écrites dans le fichier Excel {excel_file_path} avec succès.")

    def recuperer_valeurs(self):
        """
        Fonction qui récupère les valeurs des fichiers générés Excel et les retourne
        :return: Les valeurs des fichiers générés Excel
        """
        print("Récupération des valeurs...")
        valeurs_globales = []
        for fichier in os.listdir(self.dossier):
            if fichier.startswith("S") and fichier.endswith(".xlsx"):
                chemin_fichier = os.path.join(self.dossier, fichier)
                valeurs_fichier = []

                wb = load_workbook(chemin_fichier, read_only=True)

                for onglet in wb.sheetnames[1:]:
                    recup = wb[onglet]

                    valeur_case = recup.cell(row=2, column=2).value

                    valeurs_onglet = []
                    i = 56
                    while True:
                        valeur_case_1 = [recup.cell(row=i, column=1).value,
                                         recup.cell(row=i, column=12).value,
                                         recup.cell(row=i, column=13).value,
                                         recup.cell(row=i, column=14).value,
                                         recup.cell(row=i, column=15).value]
                        if any(valeur_case_1):
                            valeurs_onglet.append(valeur_case_1)
                            i += 1
                        else:
                            break

                    valeur_case_2 = [recup.cell(row=65, column=1).value,
                                     recup.cell(row=65, column=12).value,
                                     recup.cell(row=65, column=13).value,
                                     recup.cell(row=65, column=14).value,
                                     recup.cell(row=65, column=15).value]
                    valeurs_onglet.append(valeur_case_2)

                    valeurs_fichier.append((fichier, onglet, valeur_case, valeurs_onglet))

                valeurs_globales.append(valeurs_fichier)
        print("Récupération des valeurs terminée.")
        return valeurs_globales

    def inserer_dans_bd(self, valeurs_recuperes):
        """
        Fonction qui insère les valeurs récupérées dans la base de données des fichiers générés
        :param valeurs_recuperes: Les valeurs récupérées
        """
        print("Insertion des valeurs dans la base de données...")
        conn = sqlite3.connect('database/database.db')
        cursor = conn.cursor()

        # Réinitialiser la table pour insérer de nouvelles données
        cursor.execute("DELETE FROM FichiersGenere")

        valeurs_par_professeur = {}

        for valeurs_fichier in valeurs_recuperes:
            for fichier, onglet, _, valeurs_onglet in valeurs_fichier:
                for ligne in valeurs_onglet:
                    professeur = ligne[0]
                    if professeur not in valeurs_par_professeur:
                        valeurs_par_professeur[professeur] = []
                    valeurs_par_professeur[professeur].append((fichier, onglet, ligne))

        for professeur, valeurs in valeurs_par_professeur.items():
            for fichier, onglet, ligne in valeurs:
                nom_fichier_sans_extension = os.path.splitext(fichier)[0]
                cursor.execute(
                    "INSERT INTO FichiersGenere (Semestre, Prof, Ressources, H_CM, H_TD, H_TP) VALUES (?, ?, ?, ?, ?, ?)",
                    (nom_fichier_sans_extension, professeur, onglet, ligne[1], ligne[2], ligne[3]))

        conn.commit()
        conn.close()
        print("Insertion des valeurs dans la base de données terminée.")


if __name__ == "__main__":
    ScribeFileProf().run()
    print("Données écrites dans le fichier Excel avec succès.")
