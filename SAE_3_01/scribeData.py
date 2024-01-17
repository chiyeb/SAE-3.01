import sqlite3
import openpyxl as openpyxl
import os
import pandas as pd
from datetime import datetime

from insertData import insertData


class scribeData:
    instance = None

    def __new__(cls):
        if cls.instance is None:
            cls.instance = super(scribeData, cls).__new__(cls)
            cls.instance._setup()
        return cls.instance

    def _setup(self):
        """
                "Setup" l'objet : initialise la connexion à la BD
                :return:
                """
        self.conn = sqlite3.connect('database/database.db')
        self.cursor = self.conn.cursor()

    def scribeRessource(self, semestre):
        """
        Fonction qui écrit les informations de chaque ressources, dans le fichier final
        :param semestre:
        :return:
        """
        try:
            print("Début de la méthode scribeRessource")

            # On récupère les semestres uniques du planning
            self.cursor.execute("SELECT DISTINCT Semestre FROM Planning WHERE Semestre = ?", (semestre,))
            semesters = [row[0] for row in self.cursor.fetchall()]
            print(f"Semestres trouvés : {semesters}")

            # On ouvre le fichier "fichier_vierge.xlsx"
            try:
                fichier_vierge = openpyxl.load_workbook("fichiers necessaires/fichier_vierge.xlsx", read_only=False)
            except FileNotFoundError:
                print("Le fichier fichier_vierge.xlsx n'a pas été trouvé.")
                exit()
            for semester in semesters:
                print(f"Traitement du semestre : {semester}")

                # On récupère des ressources pour le semestre actuel
                self.cursor.execute("SELECT DISTINCT Ressource FROM Planning WHERE Semestre = ?", (semester,))
                resources = [row[0] for row in self.cursor.fetchall()]
                print(f"Ressources trouvées : {resources}")

                for resource in resources:
                    print(f"Traitement de la ressource : {resource}")

                    base_sheet = fichier_vierge["Feuil1"]

                    # Création d'une nouvelle feuille Excel avec le nom de la ressource
                    worksheet = fichier_vierge.copy_worksheet(base_sheet)
                    resource_key = resource.split()[0]
                    worksheet.title = resource_key

                    # On récupère des données de la base de données (Planning)
                    self.cursor.execute(
                        "SELECT Ressource, H_CM, H_TD, H_TP, COALESCE(Prof.NomProf, Planning.Resp) FROM Planning "
                        "LEFT JOIN Prof ON Planning.Resp = Prof.Acronyme "
                        "WHERE Ressource = ? AND Semestre = ?",
                        (resource, semester))
                    data_from_database = self.cursor.fetchall()

                    # On écrit les données de la base de données dans la nouvelle feuille Excel
                    if data_from_database:
                        print(f"Traitement des données ({resource}, {semester}): {data_from_database}")
                        worksheet.cell(row=5, column=2, value=data_from_database[0][1])
                        worksheet.cell(row=5, column=3, value=data_from_database[0][2])
                        worksheet.cell(row=5, column=4, value=data_from_database[0][3])
                        worksheet.cell(row=2, column=2, value=resource)
                        worksheet.cell(row=2, column=7, value=data_from_database[0][4])

                    # On récupère les données de la base de données HoraireProf
                    resource_key = resource.split()[0]
                    self.cursor.execute(
                        "SELECT Intervenant, TD, TP_Dedoubles,TP_Non_Dedoubles FROM HoraireProf WHERE ressource LIKE ? "
                        "AND TD IS NOT NULL AND TP_dedoubles IS NOT NULL AND TP_non_dedoubles IS NOT NULL",
                        (resource_key + '%',))
                    data_from_database = self.cursor.fetchall()

                    # On écrit les données dans la feuille Excel
                    if data_from_database:
                        print(f"Traitement des données ({resource}): {data_from_database}")

                        row_index0 = 8
                        row_index1 = 18
                        row_index2 = 23
                        row_index3 = 28

                        for intervenant in data_from_database:
                            # ceci permet d'écrire les intervenant
                            worksheet.cell(row=row_index0, column=1, value=intervenant[0])
                            row_index0 += 1
                            # ceci permet d'écrire les intervenant ainsi que leurs nombre de TD
                            worksheet.cell(row=row_index1, column=1, value=intervenant[0])
                            worksheet.cell(row=row_index1, column=2, value=intervenant[1])
                            row_index1 += 1
                            # ceci permet d'écrire les intervenant ainsi que leurs nombre de TD dédoublés
                            worksheet.cell(row=row_index2, column=1, value=intervenant[0])
                            worksheet.cell(row=row_index2, column=2, value=intervenant[2])
                            row_index2 += 1
                            # ceci permet d'écrire les intervenant ainsi que leurs nombre de TD non dédoublés
                            worksheet.cell(row=row_index3, column=1, value=intervenant[0])
                            worksheet.cell(row=row_index3, column=2, value=intervenant[3])
                            row_index3 += 1

                    # On récupère les noms des intervenants ayant un chiffre dans la colonne CM pour afficher les
                    # profs qui font les CM
                    resource_key = resource.split()[0]
                    self.cursor.execute(
                        "SELECT Intervenant, CM FROM HoraireProf WHERE Ressource LIKE ? AND CM IS NOT NULL",
                        (resource_key + '%',))
                    data_from_database = self.cursor.fetchall()

                    # On écrit les données dans la feuille Excel
                    if data_from_database:
                        print(f"Traitement des données ({resource}): {data_from_database}")
                        row_index = 15
                        for intervenant in data_from_database:
                            worksheet.cell(row=row_index, column=1, value=intervenant[0])
                            row_index += 1

                    self.cursor.execute(
                        "SELECT Date_Cours, Type_Cours FROM Cours WHERE Semestre = ? AND ? LIKE Ressource || '%'",
                        (semester, resource))
                    data_from_database = self.cursor.fetchall()

                    if data_from_database:
                        print(
                            f"Données de la base de données Cours ({resource}, {semester}): {data_from_database}")
                        date_type_rows = {}
                        dict_date_cours = {}
                        for i, cours_actuel in enumerate(data_from_database):
                            date_actuelle = cours_actuel[0]
                            type_cours = cours_actuel[1]
                            if date_actuelle in dict_date_cours:
                                dict_date_cours[date_actuelle].append(type_cours)
                            # Ajouter le type de cours à la liste s'il n'est pas déjà présent
                            if type_cours not in dict_date_cours.get(date_actuelle, []):
                                dict_date_cours.setdefault(date_actuelle, []).append(type_cours)
                            # Parcourir les éléments suivants pour vérifier les doublons
                        print(dict_date_cours)
                        for cle, types_cours in dict_date_cours.items():
                            # Vérifier si la date a déjà été traitée
                            if cle not in date_type_rows:
                                # Si la date n'a pas encore été traitée, obtenir un nouvel index de ligne
                                row_index5 = max(date_type_rows.values(), default=34) + 1
                                date_type_rows[cle] = row_index5
                            else:
                                row_index5 = date_type_rows[cle]
                            # Écrire la date dans la première colonne de la ligne identifiée
                            worksheet.cell(row=row_index5, column=1, value=cle)
                            t_cours_traite = {}
                            for type_cours in types_cours:
                                # Vérifie si le type de cours existe déjà dans le dictionnaire
                                if type_cours in t_cours_traite:
                                    # Augmente le compteur existant pour ce type de cours
                                    t_cours_traite[type_cours] += 2
                                else:
                                    # Initialise le compteur pour ce type de cours
                                    t_cours_traite[type_cours] = 2
                            for type_cours in set(types_cours):
                                # Concaténer le type de cours avec son compteur
                                type_cours_et_nombre = f"{type_cours} {t_cours_traite[type_cours]}H"

                                # Choix du column_index en fonction du type de cours
                                if type_cours == 'Amphi':
                                    column_index = 2
                                elif type_cours == 'TD':
                                    column_index = 3
                                elif type_cours == 'TP':
                                    column_index = 4
                                elif type_cours == 'Test':
                                    column_index = 5
                                else:
                                    continue  # Si le type de cours n'est pas reconnu, passer au suivant

                                # Écrire les données dans la feuille Excel
                                worksheet.cell(row=row_index5, column=column_index, value=type_cours_et_nombre)

                    # On récupère les données de la base de données Cours pour multiplier les valeurs de HorraireProf
                    self.cursor.execute(
                        "SELECT Type_Cours, COUNT(*) * 2 FROM Cours WHERE Semestre = ? AND ? LIKE Ressource || '%' GROUP BY Type_Cours ORDER BY Type_Cours ASC",
                        (semester, resource)
                    )
                    multiplication_values = self.cursor.fetchall()

                    if multiplication_values:
                        print(f"Traitement des données Cours ({resource}) ({semester}): {multiplication_values}")

                    # On récupère les données de la base de données HoraireProf
                    resource_key = resource.split()[0]
                    self.cursor.execute(
                        "SELECT Intervenant, CM, TD, TP_Dedoubles, TP_Non_Dedoubles, Test FROM HoraireProf WHERE ressource LIKE ?",
                        (resource_key + '%',))
                    data_from_database = self.cursor.fetchall()

                    # On récupère les données de la base de données Prof pour verfier le nom des profs qui sont dans
                    # HoraireProf
                    self.cursor.execute(
                        "SELECT NomProf FROM Prof")
                    prof_data = [row[0].upper() for row in self.cursor.fetchall()]  # Convertir en majuscules

                    if prof_data:
                        print(f"Traitement des données Prof : {prof_data}")

                    # On écrit les données dans la feuille Excel
                    if data_from_database:
                        print(f"Traitement des données GroupeProf ({resource}): {data_from_database}")

                    row_index0 = 56
                    row_index1 = 56
                    row_index2 = 56
                    row_index3 = 56
                    row_index4 = 56
                    row_index5 = 56

                    row_index6 = 65
                    row_index7 = 65
                    row_index8 = 65
                    row_index9 = 65
                    row_index10 = 65
                    row_index11 = 65

                    for intervenant in data_from_database:

                        # Multiplication des valeurs de CM, TD, TP, Test avec les valeurs d'Amphi, TD, TP, Test
                        cm_multiplier = next(
                            (count for type_cours, count in multiplication_values if type_cours == 'Amphi'), 1)
                        td_multiplier = next(
                            (count for type_cours, count in multiplication_values if type_cours == 'TD'), 1)
                        tp_multiplier = next(
                            (count for type_cours, count in multiplication_values if type_cours == 'TP'), 1)
                        test_multiplier = next(
                            (count for type_cours, count in multiplication_values if type_cours == 'Test'), 1)

                        # Vérifier si le nom de l'intervenant est dans la liste des professeurs
                        intervenant_name = intervenant[0].upper()

                        if intervenant_name in prof_data:
                            print(f"{intervenant_name} est titulaire")
                            # Écrire le nom du titulaire dans la colonne 1, ligne 65
                            worksheet.cell(row_index6, column=1, value=intervenant_name)
                            row_index6 += 1

                            # écrire les groupes de  CM dans la colonne 12, ligne 65
                            if intervenant[1] is not None:
                                worksheet.cell(row_index7, column=12, value=intervenant[1] * cm_multiplier)
                                row_index7 += 1

                            # écrire les groupes de  TD dans la colonne 13, ligne 65
                            if intervenant[2] is not None:
                                worksheet.cell(row_index8, column=13, value=intervenant[2] * td_multiplier)
                                row_index8 += 1

                            # écrire les groupes de  TP dédoublés dans la colonne 14, ligne 65
                            if intervenant[3] is not None:
                                worksheet.cell(row_index9, column=14, value=intervenant[3] * tp_multiplier)
                                row_index9 += 1

                            # écrire les groupes de  TP non dédoublésdans la colonne 15, ligne 65
                            if intervenant[4] is not None:
                                worksheet.cell(row_index10, column=15, value=intervenant[4] * tp_multiplier)
                                row_index10 += 1



                        elif intervenant_name not in prof_data:

                            print(f"{intervenant_name} est vacataire")

                            # Écrire le nom du vacataire dans la colonne 1, ligne 56
                            worksheet.cell(row_index0, column=1, value=intervenant_name)
                            row_index0 += 1

                            # écrire les groupes de  CM dans la colonne 12, ligne 56
                            if intervenant[1]:
                                worksheet.cell(row_index1, column=12, value=intervenant[1] * cm_multiplier)

                                row_index1 += 1

                            # écrire les groupes de  TD dans la colonne 13, ligne 56
                            if intervenant[2] is not None:
                                worksheet.cell(row_index2, column=13, value=intervenant[2] * td_multiplier)

                                row_index2 += 1

                            # écrire les groupes de  TP dédoublés dans la colonne 14, ligne 56
                            if intervenant[3] is not None:
                                worksheet.cell(row_index3, column=14, value=intervenant[3] * tp_multiplier)

                                row_index3 += 1

                            # écrire les groupes de  TP non dédoublés dans la colonne 15, ligne 56
                            if intervenant[4] is not None:
                                worksheet.cell(row_index4, column=15, value=intervenant[4] * tp_multiplier)

                                row_index4 += 1

                filename = f"{semestre}.xlsx"
                save_directory = 'fichiers genere'
                if not os.path.exists(save_directory):
                    os.makedirs(save_directory)
                save_path = os.path.join(save_directory, filename)
                # Sauvegarde du fichier Excel
                fichier_vierge.save(save_path)
                print(f"Les données ont été ajoutées à {save_path}")

        except sqlite3.Error as e:
            print(f"Erreur lors de la connexion à la base de données : {e}")

    def scribeHoraireTotalProf(self):
        """
        Fonction qui écrit les heures de chaque professeurs pour chaque type de cours dans un fichier Excel.
        :return:
        """
        try:
            insertdata_instance = insertData()
            insertdata_instance.insertNombreHeureProf()
            # récupération du nombres d'heures de chaque profs
            self.cursor.execute("SELECT * FROM HoraireTotalProf")
            heures = self.cursor.fetchall()
            colones = [description[0] for description in self.cursor.description]
            excel = pd.DataFrame(heures, columns=colones)
            # calcul du nombre total d'heure pour chaque prof (somme de toutes les heures)
            excel['Total_Hours'] = excel[['H_CM', 'H_TD', 'H_TP_D', 'H_TP_ND', 'H_TEST']].sum(axis=1)
            # on renomme les colonnes du fichier excel
            excel.rename(columns={
                'Prof': 'Nom du prof',
                'H_CM': 'Nombre d\'heure de CM',
                'H_TD': 'Nombre d\'heure de TD',
                'H_TP_D': 'Nombre d\'heure de TP Dedoubles',
                'H_TP_ND': 'Nombre d\'heure de TP Non Dedoubles',
                'H_TEST': 'Nombre d\'heure de test',
                'Total_Hours': 'Nombre d\'heure total'
            }, inplace=True)

            dossier = 'fichiers genere'
            if not os.path.exists(dossier):
                os.makedirs(dossier)
            fichier = "Professeurs_Horaires.xlsx"
            lieu_enregistrement = os.path.join(dossier, fichier)
            excel.to_excel(lieu_enregistrement, index=False)
            print(f"Les données ont été écrites dans le fichier {lieu_enregistrement}")

        except sqlite3.Error as e:
            print(f"Erreur lors de la récupération des données : {e}")

# test = scribeData()
# test.scribeHoraireTotalProf()
