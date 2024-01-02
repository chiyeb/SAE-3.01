import sqlite3
import openpyxl
from openpyxl.styles import Font

# Se connecter à la base de données
db = sqlite3.connect("database/database.db")
db.row_factory = sqlite3.Row # Permet d'accéder aux attributs par leur nom, et non par leur indice

# Créer un nouveau fichier Excel
fichier = openpyxl.Workbook()

# Sélectionner l'onglet
onglet = fichier.active

# Nommer l'onglet
onglet.title = "Heures par Prof"

# Fusionner les cellules A1 et A2
onglet.merge_cells("A1:A2")

# Fusionner les cellules B1 et B2
onglet.merge_cells("B1:B2")

# Fusionner les cellules C1 et C2
onglet.merge_cells("C1:C2")

# Fusionner les cellules D1 et D2
onglet.merge_cells("D1:D2")

# Fusionner les cellules E1, F1, G1, H1, I1, et J1
onglet.merge_cells("E1:J1")

# Affecter un nom à chaque colonne
data = {
    "A1": "Nom",
    "B1": "BUT 1 / BUT 2 / BUT 3",
    "C1": "Parcours A / Parcours B",
    "D1": "FI / FA",
    "E1": "Nombre d'heures prévues",
    "E2": "CM",
    "F2": "TD",
    "G2": "TP\n(1/2 groupe)",
    "H2": "TP\n(non dédoublé)",
    "I2": "TP à déclarer ARES",
    "J2": "Total en HETD"
}

# Insérer chaque valeur de "data" dans l'onglet
for cellule, valeur in data.items():
    onglet[cellule] = valeur
    onglet[cellule].font = Font(bold=True) # Met ces valeurs en gras

# Récupérer toutes les données de la table HoraireProf
curseur = db.execute("SELECT * FROM HoraireProf").fetchall()

for tuple in curseur:
    if tuple["Ressource"][1] == "1" or tuple["Ressource"] == "2":
        année = "BUT1"
    elif tuple["Ressource"][1] == "3" or tuple["Ressource"] == "4":
        année = "BUT2"
    elif tuple["Ressource"][1] == "5" or tuple["Ressource"] == "6":
        année = "BUT3"
    onglet.append([tuple["intervenant"], année])

# Sauvegarder le classeur
fichier.save("fichiers genere/Heures par prof.xlsx")

