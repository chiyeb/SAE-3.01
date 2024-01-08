import sqlite3
import math
from datetime import datetime

import openpyxl
import pandas as pd

from insertData import insertData
from selectFile import *


class recupData:
    """
    Classe permettant de récupérer certaines données dans les fichiers
    """
    instance = None
    files = None

    def __new__(cls):
        if cls.instance is None:
            cls.instance = super(recupData, cls).__new__(cls)
            cls.instance._setup()
        return cls.instance

    def _setup(self):
        """
        "Setup" l'objet : initialise la connexion à la BD
        :return:
        """
        self.files = selectFile()
        # Initialise la connexion à la base de données
        self.conn = sqlite3.connect('database/database.db')
        self.cursor = self.conn.cursor()

    def trouverVal(self, semestre, semestre_onglet):
        """
        Fonction qui trouve les ressources dans le fichier planning, récupère les valeurs des heures de CM,TD et TP.
        :param semestre:
        :param semestre_onglet:
        :return:
        """
        planning = pd.ExcelFile(self.files.planning_file)
        self.cursor.execute("SELECT Libelle, Num_Res FROM Maquette WHERE Semestre = ?", (semestre,))
        resultats = self.cursor.fetchall()
        print(resultats)
        for row in resultats:
            num_res = row[1]
            libelle = row[0]
            S = pd.read_excel(planning, semestre_onglet)
            for index, row in S.iterrows():
                if num_res in row.values:
                    index_num_res = row.values.tolist().index(num_res)
                    valeur_case_3 = 0 if pd.isna(row.iloc[index_num_res + 3]) else row.iloc[index_num_res + 3]
                    valeur_case_5 = 0 if pd.isna(row.iloc[index_num_res + 5]) else row.iloc[index_num_res + 5]
                    valeur_case_7 = 0 if pd.isna(row.iloc[index_num_res + 7]) else row.iloc[index_num_res + 7]
                    valeur_case_10 = row.iloc[index_num_res + 10]
                    additions = 0
                    if (isinstance(valeur_case_3, int) and isinstance(valeur_case_7, int)
                            and isinstance(valeur_case_5, int)):
                        additions = valeur_case_3 + valeur_case_5 + valeur_case_7
                        if index_num_res + 10 < len(row) and additions > 0:
                            print(f"Num_Res trouvé: {num_res}")
                            print(f"3e case après le mot: {valeur_case_3}")
                            print(f"5e case: {valeur_case_5}")
                            print(f"7e case: {valeur_case_7}")
                            print(f"10e case: {valeur_case_10}")
                            insertdata = insertData()
                            insertdata.insert_planning(semestre, libelle, valeur_case_3, valeur_case_5, valeur_case_7,
                                                       valeur_case_10)

    def onglet_existe(self, fichier, nom_onglet):
        """
        Vérifie si l'onglet existe
        :param fichier:
        :param nom_onglet:
        :return:
        """
        wb = openpyxl.load_workbook(fichier, read_only=True)
        # Retourne True si l'onglet existe, False sinon
        return nom_onglet in wb.sheetnames

    def recupHProf(self, semestre, semestre_onglet):
        """
        Récupère les heures de chaque professeur
        :param semestre:
        :param semestre_onglet:
        :return:
        """
        # Charge le fichier Excel
        fichier = self.files.QFQ_file
        if self.onglet_existe(fichier, semestre_onglet):
            df = pd.read_excel(fichier, semestre_onglet)
            donnees = {}
            ressourceActuelle = None

            for _, row in df.iterrows():
                resource = row['Ressource']
                intervenant = row['Intervenants']
                cm = row['CM']
                td = row['TD']
                tp_non_dedoubles = row['TP (non dédoublés)']
                tp_dedoubles = row['TP (dédoublés)']
                test = row['Test']

                # Vérifie si une nouvelle ressource commence
                if pd.notna(resource):
                    ressourceActuelle = resource
                    donnees[ressourceActuelle] = {}

                # Vérifie si la ligne contient des données d'intervenant
                if pd.notna(intervenant):
                    if ressourceActuelle not in donnees:
                        donnees[ressourceActuelle] = {}
                    if intervenant not in donnees[ressourceActuelle]:
                        donnees[ressourceActuelle][intervenant] = []

                    donnees[ressourceActuelle][intervenant].append({
                        'CM': cm,
                        'TD': td,
                        'TP (non dédoublés)': tp_non_dedoubles,
                        'TP (dédoublés)': tp_dedoubles,
                        'Test': test})

            # Insère ou met à jour les enregistrements dans la base de données
            for resource, intervenant_data in donnees.items():
                if resource not in donnees:
                    donnees[resource] = {
                        None: [{}]}  # Insère un enregistrement avec des valeurs nulles si la ressource n'est pas présente

                for intervenant, data_list in intervenant_data.items():
                    print(f"Ressource: {resource} - Intervenant : {intervenant} ")

                    if data_list:
                        for d in data_list:
                            # Vérifie si l'enregistrement existe déjà dans la base de données
                            self.cursor.execute(
                                "SELECT * FROM HoraireProf WHERE Semestre = ? AND Ressource = ? AND Intervenant = ?",
                                (semestre, resource, intervenant))
                            existing_record = self.cursor.fetchone()

                            if existing_record:
                                # Met à jour l'enregistrement existant
                                self.cursor.execute(
                                    "UPDATE HoraireProf SET CM = ?, TD = ?, TP_Non_Dedoubles = ?, TP_Dedoubles = ?, "
                                    "Test = ? WHERE Semestre = ? AND Ressource = ? AND Intervenant = ?",
                                    (d['CM'], d['TD'], d['TP (non dédoublés)'], d['TP (dédoublés)'], d['Test'],
                                     semestre, resource, intervenant))
                            else:
                                # Insère un nouvel enregistrement
                                self.cursor.execute(
                                    "INSERT INTO HoraireProf (Semestre, Ressource, Intervenant, CM, TD, TP_non_dedoubles, "
                                    "TP_Dedoubles, Test) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                                    (semestre, resource, intervenant, d['CM'], d['TD'], d['TP (non dédoublés)'],
                                     d['TP (dédoublés)'],
                                     d['Test']))

                            self.conn.commit()
                            # Différents affichage console
                            print(f"    - CM : {d['CM']} heure" if pd.notna(d['CM']) else "    - CM : Non spécifié")
                            print(f"    - TD : {d['TD']} heure" if pd.notna(d['TD']) else "    - TD : Non spécifié")
                            print(f"    - TP (non dédoublés) : {d['TP (non dédoublés)']} heure" if pd.notna(
                                d['TP (non dédoublés)']) else "    - TP (non dédoublés) : Non spécifié")
                            print(f"    - TP (dédoublés) : {d['TP (dédoublés)']} heure" if pd.notna(
                                d['TP (dédoublés)']) else "    - TP (dédoublés) : Non spécifié")
                            print(f"    - Test : {d['Test']} heure" if pd.notna(d['Test']) else "    - Test : Non spécifié")
                            print("\n")
                            print(f"Extraction pour la ressource: {resource} terminée !")
                    else:
                        print("Aucune données pour cette ressource.")
                        print("\n")

    def recupNomProf(self):
        """
        Récupère les noms des professeurs dans le fichier texte, ainsi que leur acronymes
        :return:
        """
        # Ouvre le fichier en mode lecture
        with open(self.files.nom_prof_file, "r") as fichier:
            # Lire chaque ligne du fichier
            for ligne in fichier:
                # Divise la ligne en utilisant le signe "=" comme séparateur
                parties = ligne.strip().split("=")

                # Vérifie si la ligne a été correctement divisée en deux parties
                if len(parties) == 2:
                    # Exécute la requête SQL
                    resultat_requete = self.cursor.execute("SELECT Acronyme, NomProf FROM PROF WHERE Acronyme = ?",
                                                           (parties[0],)).fetchall()
                    # Vérifie si la requête a renvoyé des résultats, pour savoir si le prof existe déjà
                    if resultat_requete:
                        # On update le nom du prof
                        self.cursor.execute("UPDATE PROF SET NomProf = ? WHERE Acronyme = ?", (parties[1], parties[0]))
                        self.conn.commit()
                    else:
                        # On insère le nouveau prof
                        self.cursor.execute(
                            "INSERT INTO Prof (Acronyme, NomProf) VALUES (?, ?)",
                            (parties[0], parties[1],))
                        # sauvegarder dans la base de données
                        self.conn.commit()
                else:
                    print(f"Erreur de format sur la ligne : {ligne}")

    def recupRCouleur(self, semestre, semestre_onglet):
        """
        Fonction pour récupérer les couleurs lié à chaque ressources pour un semestre précis
        :param semestre:
        :param semestre_onglet:
        :return:
        """
        ressourceCouleur = {}
        self.cursor.execute("SELECT Num_Res FROM Maquette WHERE Semestre = ?", (semestre,))
        resultats = [item[0] for item in self.cursor.fetchall()]
        # On ouvre le fichier excel
        fichier = openpyxl.load_workbook(self.files.planning_file)
        fichierOngletSemestre = fichier[semestre_onglet]
        for row in fichierOngletSemestre.iter_rows():
            for cell in row:
                # Si ressource trouvé
                if cell.value in resultats:
                    couleur = cell.fill.start_color.rgb
                    # Si couleur est autre que blanche
                    if couleur != '00000000' and couleur:
                        # On récupère la couleur de la cellule
                        ressourceCouleur[cell.value] = couleur
        return ressourceCouleur

    def trouverTypeCours2eRange(self, semestre_onglet):
        """
        Récupère la premiere occurence de chaque type de cours dans le fichier planning
        :param semestre_onglet:
        :return:
        """
        fichier = openpyxl.load_workbook(self.files.planning_file, data_only=True)
        fichierOngletSemestre = fichier[semestre_onglet]
        type_cours_dict = {}
        for row in fichierOngletSemestre.iter_rows(min_row=1, max_row=2):
            # On itère chaque cellule
            for cell in row:
                # Si la valeur de la celulle = ["Cours", "TD", "TP", "Test"]
                if cell.value in ["Cours", "TD", "TP", "Test"]:
                    # On récupère la valeur de la colone de la cellule
                    type_cours_dict[cell.value] = cell.column

        print(type_cours_dict)
        return type_cours_dict

    def trouverTypeCours1erRange(self, semestre_onglet):
        """
        Récupère la deuxième occurence de chaque type de cours dans le fichier planning
        :param semestre_onglet:
        :return:
        """
        fichier = openpyxl.load_workbook(self.files.planning_file, data_only=True)
        fichierOngletSemestre = fichier[semestre_onglet]
        type_cours_dict = {}
        # Les valeurs à trouver
        valeurs_a_trouver = {"Cours", "TD", "TP", "Test"}
        for row in fichierOngletSemestre.iter_rows(min_row=1, max_row=2):
            # On itère chaque cellule
            for cell in row:
                # Si la valeur de la celulle = ["Cours", "TD", "TP", "Test"]
                if cell.value in ["Cours", "TD", "TP", "Test"]:
                    # On récupère la valeur de la colonne de la cellule
                    type_cours_dict[cell.value] = cell.column
                    # On enlève de "valeur_a_trouver" la valeur de la cellule
                    valeurs_a_trouver.remove(cell.value)
                # S'il y à plus de valeur dans "valeur_a_trouver"
                if not valeurs_a_trouver:
                    return type_cours_dict
        print(type_cours_dict)
        return type_cours_dict

    def recupXetY(self, semestre, semestre_onglet):
        """
        Récupère chaque cours dans le fichier planning
        :param semestre:
        :param semestre_onglet:
        :return:
        """
        # On réinitialise les valeurs de la base de donnée
        self.cursor.execute("UPDATE Horaires SET NbCours = 0 WHERE Semestre = ?",
                            (semestre,))
        # Récupération du fichier planning
        fichier = openpyxl.load_workbook(self.files.planning_file, data_only=True)
        # Appel des fonctions nécéssaire
        ressourceCouleur = self.recupRCouleur(semestre, semestre_onglet)
        fichierOngletSemestre = fichier[semestre_onglet]
        type_cours_dict1 = self.trouverTypeCours1erRange(semestre_onglet)
        type_cours_dict2 = self.trouverTypeCours2eRange(semestre_onglet)
        for col in fichierOngletSemestre.iter_cols():
            if col[0].value is not None and 'Date' in col[0].value:
                # Parcourir les lignes pour la colonne de date
                for cell in col:
                    cell_value = cell.value
                    # Si c'est un date
                    if isinstance(cell_value, datetime):
                        date = cell_value.date()
                    else:
                        date = cell_value
                    # s'il y a une date trouvé
                    if date:
                        row_index = cell.row
                        for row_cell in fichierOngletSemestre[row_index]:
                            # Récupérer la valeur et la couleur de chaque cellule de la ligne
                            if row_cell.value == "X" or row_cell.value == "Y":
                                if row_cell.value == "X":
                                    salle = "TD/Amphi"
                                else:
                                    salle = "Machine"
                                # Appel de la fonction typeCours qui permet de vérifier quel type de cours est une case
                                tCours = self.typeCours(type_cours_dict1, type_cours_dict2, row_cell.column)
                                valeur = row_cell.value
                                # Récupération de la couleur de la cellule
                                couleur = row_cell.fill.start_color.rgb
                                # On récupère les coordonnées du cours (X,Y)
                                idCours = row_cell.coordinate
                                print(date)
                                for cle, valeur in ressourceCouleur.items():
                                    # Si la couleur de la cellule = la couleur d'une ressource
                                    if valeur == couleur:
                                        # On vérifie si le cours existe déjà dans la BD
                                        self.cursor.execute("SELECT IdCours, Semestre FROM Cours WHERE IdCours = ? "
                                                            "AND Semestre = ?", (idCours, semestre))
                                        resultCours = self.cursor.fetchone()
                                        # On vérifie si la ressource et le type de cours existe déjà dans la BD
                                        self.cursor.execute(
                                            "SELECT Ressource, Type_Cours FROM Horaires WHERE Ressource "
                                            "= ? AND Type_Cours = ?",
                                            (cle, tCours))
                                        resultHoraires = self.cursor.fetchone()
                                        # Si le cours n'existe pas
                                        if not resultCours:
                                            # Si un commentaire existe pour la cellule
                                            if row_cell.comment:
                                                # On récupère le commentaire et l'insère
                                                commentaire = row_cell.comment.text
                                                self.cursor.execute("INSERT INTO Cours (IdCours, Semestre, Ressource, "
                                                                    "Date_Cours, Commentaire, Type_Cours, Salle) VALUES (?, "
                                                                    "?, ?, ?, ?, ?, ?)",
                                                                    (idCours, semestre, cle, date, commentaire, tCours,
                                                                     salle))
                                            # Si pas de commentaires dans la cellule
                                            else:
                                                self.cursor.execute("INSERT INTO Cours (IdCours, Semestre, Ressource, "
                                                                    "Date_Cours, Type_Cours, Salle) VALUES (?, ?, "
                                                                    "?, ?, ?, ?)",
                                                                    (idCours, semestre, cle, date, tCours, salle))
                                            self.conn.commit()
                                        # Si le cours n'existe pas
                                        else:
                                            if row_cell.comment:
                                                commentaire = row_cell.comment.text
                                                self.cursor.execute("UPDATE Cours SET Ressource = ?, Date_Cours = ?, "
                                                                    "Commentaire = ?, Type_Cours = ?, Salle = ? WHERE IdCours = ? ",
                                                                    (cle, date, commentaire, tCours, salle, idCours,))
                                            else:
                                                self.cursor.execute("UPDATE Cours SET Ressource = ?, Date_Cours = ?, "
                                                                    "Type_Cours = ?, Salle = ? WHERE IdCours = ? ",
                                                                    (cle, date, tCours, salle, idCours))
                                            self.conn.commit()
                                        # Si le type de cours existe déjà pour le semestre et la ressource dans la BD
                                        if resultHoraires:
                                            self.cursor.execute("UPDATE Horaires SET NbCours = NbCours+2 WHERE "
                                                                "Ressource = ? AND Type_Cours = ?",
                                                                (cle, tCours))
                                        else:
                                            self.cursor.execute("INSERT INTO Horaires (Semestre, Ressource, "
                                                                "Type_Cours, nbCours, Salle) VALUES (?, ?, ?, 2, ?)",
                                                                (semestre, cle, tCours, salle))
                                        self.cursor.connection.commit()
                                        print(f"Cle: {cle} Valeur: {valeur}, Couleur: {couleur}")

    def typeCours(self, type_cours_dict1, type_cours_dict2, col):
        """
        Fonction qui vérifie pour une case précise, quel type de cours c'est
        :param type_cours_dict1:
        :param type_cours_dict2:
        :param col:
        :return:
        """
        tCours = None
        if type_cours_dict1["Cours"] <= col <= type_cours_dict1["TD"]:
            tCours = "Amphi"
        if type_cours_dict1["TD"] <= col <= type_cours_dict1["TP"]:
            tCours = "TD"
        if type_cours_dict1["TP"] <= col <= type_cours_dict1["Test"]:
            tCours = "TP"
        if type_cours_dict1["Test"] <= col <= type_cours_dict2["Cours"]:
            tCours = "Test"
        if type_cours_dict2["Cours"] <= col <= type_cours_dict2["TD"]:
            tCours = "Amphi"
        if type_cours_dict2["TD"] <= col <= type_cours_dict2["TP"]:
            tCours = "TD"
        if type_cours_dict2["TP"] <= col <= type_cours_dict2["Test"]:
            tCours = "TP"
        if col >= type_cours_dict2["Test"]:
            tCours = "Test"
        return tCours
