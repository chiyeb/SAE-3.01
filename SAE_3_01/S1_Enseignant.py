import sqlite3
import openpyxl

# Connexion à la base de données SQLite
try:
    with sqlite3.connect("database/database.db") as conn:
        cursor = conn.cursor()

        # Exécutez une requête SQL pour récupérer toutes les ressources depuis la base de données
        cursor.execute("SELECT DISTINCT Ressource FROM Planning WHERE Semestre = 'S1'")

        # Récupérez toutes les ressources dans une liste
        resources = [row[0] for row in cursor.fetchall()]

        # Chuarger un classeur Excel existant s'il existe, sinon créer un noveau
        try:
            workbook = openpyxl.load_workbook("S1.xlsx")
        except FileNotFoundError:
            print("Le fichier S1.xlsx n'a pas été trouvé.")
            exit()

        for resource in resources:
            # Créer une nouvelle feuille avec le nom de la ressource
            worksheet = workbook.create_sheet(title=resource)

            # Copier toutes les données de la feuille de base (S2.xlsx) dans la nouvelle feuille
            base_sheet = workbook["Feuil1"]
            for row in base_sheet.iter_rows(min_row=1, max_col=base_sheet.max_column, max_row=base_sheet.max_row):
                worksheet.append([cell.value for cell in row])

            # Exécutez une requête SQL pour récupérer les données de la ressource actuelle
            cursor.execute("SELECT Ressource, H_CM, H_TD, H_TP, Resp FROM Planning WHERE Ressource = ?", (resource,))

            # Récupérez les données dans une liste de tuples
            data_from_database = cursor.fetchall()

            # Écrire les données spécifiques dans la nouvelle feuille
            worksheet.cell(row=5, column=2, value=data_from_database[0][1])
            worksheet.cell(row=5, column=3, value=data_from_database[0][2])
            worksheet.cell(row=5, column=4, value=data_from_database[0][3])
            worksheet.cell(row=2, column=2, value=resource)
            worksheet.cell(row=2, column=7, value=data_from_database[0][4])


        # Sauvegarder le classeur Excel
        workbook.save("S1.xlsx")

        # Message pour confirmer les ajouts dans le fichier Excel
        print("Les données ont été ajoutées à S1.xlsx")

# Si une erreur survient lors de la connexion à la base de données, affichez un message d'erreur
except sqlite3.Error as e:
    print(f"Erreur lors de la connexion à la base de données : {e}")
